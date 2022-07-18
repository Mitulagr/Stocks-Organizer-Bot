import requests
from bs4 import BeautifulSoup
import datetime
import json
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill, Font, colors, Alignment
from openpyxl.cell import Cell
import time


def price (stockcode) : 

    stockcode = stockcode.upper()
    if(stockcode!="NIFTY" and stockcode!="SENSEX") : 
        try : 
            stock_url  = 'https://www.nseindia.com/live_market/dynaContent/live_watch/get_quote/GetQuote.jsp?symbol='+str(stockcode)
            headers = {'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.117 Safari/537.36'}
            response = requests.get(stock_url, headers=headers)
            response
            soup = BeautifulSoup(response.text, 'html.parser')
            data_array = soup.find(id='responseDiv').getText().strip().split(":")
            type (data_array)
            for item in data_array:
                if 'lastPrice' in item:
                    index = data_array.index(item)+1
                    latestPrice=data_array[index].split('"')[1]
                    break
            return float(latestPrice.replace(',', ''))      
        except : return -1   
    elif(stockcode=="NIFTY") : 
        stock_url  = 'https://economictimes.indiatimes.com/indices/nifty_50_companies'
        headers = {'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.117 Safari/537.36'}
        response = requests.get(stock_url, headers=headers)
        soup = BeautifulSoup(response.text, 'html.parser')
        data_array = soup.find(id='ltp').getText().strip().split(":")
        return (float(data_array[0]))     
    elif(stockcode=="SENSEX") :
        stock_url  = 'https://economictimes.indiatimes.com/indices/sensex_30_companies'
        headers = {'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.117 Safari/537.36'}
        response = requests.get(stock_url, headers=headers)
        soup = BeautifulSoup(response.text, 'html.parser')
        data_array = soup.find(id='ltp').getText().strip().split(":")
        return (float(data_array[0]))  

def open_last() :
    with open('last.json', 'r') as openfile : last = json.load(openfile)
    return datetime.datetime.strptime(last, '%Y-%m-%d %H:%M:%S')

def save_last(last) :
    last = last.strftime('%Y-%m-%d %H:%M:%S')
    with open("last.json", "w") as outfile : json.dump(last, outfile)

def holiday(t) :
    if(t.weekday() in [5,6]) : return True
    d = int(t.strftime("%d"))
    m = int(t.strftime("%m"))
    y = t.strftime("%Y")
    h21 = [[],[],[],[],[],[],[],[],[19],[10],[15],[4,5,19],[]]
    if(y=='2021') : 
        if(d in h21[m]) : return True
    h22 = [[],[26],[],[1,18],[14,15],[3],[],[],[9,15,31],[],[5,24,25],[8],[]]
    if(y=='2022') : 
        if(d in h22[m]) : return True    
    return False    

def market_open(now) :
    if holiday(now) : return False     
    h = now.hour
    m = now.minute
    if (h>=10 and h<=15 and (m==30 or m==0)) or (h==9 and m==30) : return True
    return False

def slp() :
    t = datetime.datetime.now()
    while True :
        if market_open(t) :
            time.sleep((t-datetime.datetime.now()).total_seconds())
            return 0
        else : t = t + datetime.timedelta(seconds = 15)  

def add(t,nw=None) : 
    save_last(t)
    nd = t.strftime("%d-%m-%Y")
    nt = t.strftime("%I:%M %p")
    wb = load_workbook('dontopen.xlsx')
    ws1 = wb["Data"]    
    f1 = Font(name='Bahnschrift',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
    if nw is None :   
        with open('stocks.json', 'r') as openfile : stk = json.load(openfile)
        with open('wallbal.json', 'r') as openfile : wallbal = json.load(openfile)
        sum = 0 
        for i in range (len(stk)) : 
            p = price(stk[i][0])
            sum = sum + p*stk[i][1]
        nw = sum + wallbal
        nw = "{:.2f}".format(nw)  
        nw = float(nw)       
    with open('datarow.json', 'r') as openfile : rmax = json.load(openfile)
    rmax = rmax + 1
    if(rmax%13==0) :
        nwprev = ws1.cell(row = 15, column = 5).value
        f1 = Font(name='Bahnschrift',size=12,bold=False,italic=False,vertAlign=None,underline='single',strike=False,color='FF000000')
        if(nw>nwprev) : f1 = Font(name='Bahnschrift',size=12,bold=False,italic=False,vertAlign=None,underline='single',strike=False,color='FF008000')  
        if(nw<nwprev) : f1 = Font(name='Bahnschrift',size=12,bold=False,italic=False,vertAlign=None,underline='single',strike=False,color='FFFF0000') 
    with open("datarow.json", "w") as outfile : json.dump(rmax, outfile)
    ws1.move_range(f"A3:E{3+rmax}", rows=1)
    c1 = ws1.cell(row = 3, column = 1)
    c1.font = f1
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.value = nd 
    c1 = ws1.cell(row = 3, column = 3)
    c1.font = f1
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.value = nt
    c1 = ws1.cell(row = 3, column = 5)
    c1.font = f1
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.value = nw
    ws1 = wb["DayData"] 
    if(rmax%13==1) : 
        ws1.move_range(f"A3:C{4+(rmax//13)}", rows=1)
    c1 = ws1.cell(row = 3, column = 1)
    c1.font = f1
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.value = t.strftime('%Y-%m-%d %H:%M:%S')
    c1 = ws1.cell(row = 3, column = 3)
    c1.font = f1
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.value = nw
    wb.save('dontopen.xlsx')    
    try : wb.save('Stock_Chart.xlsx')
    except : pass 

def check(t) :
    last = open_last()
    mindiff = ((t-last).total_seconds())/60
    if(mindiff<31) : return True
    last = last + datetime.timedelta(minutes = 30)
    save_last(last)
    if market_open(last) :
        wb = load_workbook('dontopen.xlsx')
        ws1 = wb["Data"]    
        nw = ws1.cell(row = 3, column = 5).value
        add(last,nw)
    check(t) 


check(datetime.datetime.now())     
slp()

while True : 

    now = datetime.datetime.now()

    if market_open(now) :
        add(now)
        time.sleep(60)

    slp()
                   

