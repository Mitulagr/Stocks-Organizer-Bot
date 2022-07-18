import random
import discord
from discord import colour
import requests
from discord.ext import commands, tasks
from bs4 import BeautifulSoup
import datetime
import json
import openpyxl
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill, Font, colors, Alignment
from openpyxl.cell import Cell
import time
import S1_Chart as sc

def price (stockcode) : 

    stockcode = stockcode.upper()
    if(stockcode!="NIFTY" and stockcode!="SENSEX") : 
        try : 
            stock_url  = 'https://www.nseindia.com/live_market/dynaContent/live_watch/get_quote/GetQuote.jsp?symbol='+str(stockcode)
            headers = {'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.117 Safari/537.36'}
            response = requests.get(stock_url, headers=headers)
            response
            soup = BeautifulSoup(response.text, 'html.parser')
            data_array = soup.find(id='responseDiv').getText().strip().split(":")
            type (data_array)
            for item in data_array:
                if 'lastPrice' in item:
                    index = data_array.index(item)+1
                    latestPrice=data_array[index].split('"')[1]
                    break
            return float(latestPrice.replace(',', ''))      
        except : return -1   
    elif(stockcode=="NIFTY") : 
        stock_url  = 'https://economictimes.indiatimes.com/indices/nifty_50_companies'
        headers = {'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.117 Safari/537.36'}
        response = requests.get(stock_url, headers=headers)
        soup = BeautifulSoup(response.text, 'html.parser')
        data_array = soup.find(id='ltp').getText().strip().split(":")
        return (float(data_array[0]))     
    elif(stockcode=="SENSEX") :
        stock_url  = 'https://economictimes.indiatimes.com/indices/sensex_30_companies'
        headers = {'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.117 Safari/537.36'}
        response = requests.get(stock_url, headers=headers)
        soup = BeautifulSoup(response.text, 'html.parser')
        data_array = soup.find(id='ltp').getText().strip().split(":")
        return (float(data_array[0]))  

def update_holdings(stk) : 

    wb = load_workbook('Stock.xlsx')
    ws1 = wb["Holdings"]
    f1 = Font(name='Bahnschrift',size=11,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
    sum = 0 
    for i in range (len(stk)) : 
        p = price(stk[i][0])
        sum = sum + p*stk[i][1]
        c1 = ws1.cell(row = 5+i*2, column = 1)
        c1.font = f1
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.value = f"{i+1})"
        c1 = ws1.cell(row = 5+i*2, column = 3)
        c1.font = f1
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.value = stk[i][0]
        c1 = ws1.cell(row = 5+i*2, column = 5)
        c1.font = f1
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.value = "{:.2f}".format(p)
        c1 = ws1.cell(row = 5+i*2, column = 7)
        c1.font = f1
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.value = stk[i][1]
        c1 = ws1.cell(row = 5+i*2, column = 9)
        c1.font = f1
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.value = "{:.2f}".format(p*stk[i][1])
    f2 = Font(name='Bahnschrift',size=12,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')  
    c1 = ws1.cell(row = 5+len(stk)*2, column = 3)
    c2 = ws1.cell(row = 5+len(stk)*2, column = 9) 
    c1.font = f2
    c2.font = f2
    c1.alignment = Alignment(horizontal="center", vertical="center") 
    c2.alignment = Alignment(horizontal="center", vertical="center") 
    c1.value = "Total"
    c2.value = "{:.2f}".format(sum)
    c1 = ws1.cell(row = 7+len(stk)*2, column = 3)
    c1.value = ""
    c1 = ws1.cell(row = 7+len(stk)*2, column = 9)
    c1.value = ""
    c1 = ws1.cell(row = 5+len(stk)*2, column = 1)
    c1.value = ""
    c1 = ws1.cell(row = 5+len(stk)*2, column = 5)
    c1.value = ""
    c1 = ws1.cell(row = 5+len(stk)*2, column = 7)
    c1.value = ""
    f3 = Font(name='Bahnschrift',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000') 
    f4 = Font(name='Times New Roman',size=14,bold=True,italic=False,vertAlign=None,underline='single',strike=False,color='FF000000') 
    c1 = ws1.cell(row = 6, column = 14)
    c2 = ws1.cell(row = 8, column = 14)
    c1.font = f3
    c2.font = f4
    c1.alignment = Alignment(horizontal="center", vertical="center") 
    c2.alignment = Alignment(horizontal="center", vertical="center")  
    c1.value = ws1.cell(row = 5+len(stk)*2, column = 9).value
    c2.value = "{:.2f}".format(float(c1.value) + float(ws1.cell(row = 4, column = 14).value))
    wb.save('Stock.xlsx')

def update_wallet(diff,type) : 

    with open('walletdep.json', 'r') as openfile : w = json.load(openfile)
    if(type=="Deposit") : w = w + diff
    if(type=="Withdraw") : w = w - diff
    with open("walletdep.json", "w") as outfile : json.dump(w, outfile)
    wb = load_workbook('Stock.xlsx')
    ws1 = wb["Trades"]    
    f1 = Font(name='Bahnschrift',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
    with open('traderow.json', 'r') as openfile : rrmax = json.load(openfile)
    rmax = rrmax[1]
    rrmax[1] = rrmax[1] + 1
    with open("traderow.json", "w") as outfile : json.dump(rrmax, outfile)
    ws1.move_range(f"S4:Y{4+rmax*2}", rows=2)
    c1 = ws1.cell(row = 4, column = 19)
    c1.font = f1
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.value = datetime.date.today().strftime('%d-%m-%Y') 
    c1 = ws1.cell(row = 4, column = 21)
    c1.font = f1
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.value = type
    if(type in ["Stock Buy","Brockerage","Monthly","Withdraw"]) :
        diff = diff*-1
        f2 = Font(name='Bahnschrift',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FFFF0000')
    else : f2 = Font(name='Bahnschrift',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF008000')    
    c1 = ws1.cell(row = 4, column = 25)
    c1.font = f1
    c1.alignment = Alignment(horizontal="center", vertical="center")
    try : prev = float(ws1.cell(row = 6, column = 25).value)
    except : prev = 0.0 
    wlt = diff + prev
    c1.value = "{:.2f}".format(wlt)
    c1 = ws1.cell(row = 4, column = 23)
    c1.font = f2
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.value = "{:.2f}".format(diff)
    ws1 = wb["Holdings"]
    f1 = Font(name='Bahnschrift',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000') 
    f2 = Font(name='Times New Roman',size=14,bold=True,italic=False,vertAlign=None,underline='single',strike=False,color='FF000000') 
    c1 = ws1.cell(row = 4, column = 14)
    c2 = ws1.cell(row = 8, column = 14)
    c1.font = f1
    c2.font = f2
    c1.alignment = Alignment(horizontal="center", vertical="center") 
    c2.alignment = Alignment(horizontal="center", vertical="center")  
    c1.value = "{:.2f}".format(wlt)
    c2.value = "{:.2f}".format(float(c1.value) + float(ws1.cell(row = 6, column = 14).value))
    wb.save('Stock.xlsx')
    with open("wallbal.json", "w") as outfile : json.dump(wlt, outfile) 
 
def update_trades(stk,b,c,d,type) : 

    wb = load_workbook('Stock.xlsx')
    ws1 = wb["Trades"]    
    f1 = Font(name='Bahnschrift',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
    with open('traderow.json', 'r') as openfile : rrmax = json.load(openfile)
    rmax = rrmax[0]
    rrmax[0] = rrmax[0] + 1
    with open("traderow.json", "w") as outfile : json.dump(rrmax, outfile)
    ws1.move_range(f"A4:O{4+rmax*2}", rows=2)
    c1 = ws1.cell(row = 4, column = 1)
    c1.font = f1
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.value = datetime.date.today().strftime('%d-%m-%Y') 
    c1 = ws1.cell(row = 4, column = 3)
    c1.font = f1
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.value = b
    c1 = ws1.cell(row = 4, column = 5)
    c1.font = f1
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.value = type
    c1 = ws1.cell(row = 4, column = 7)
    c1.font = f1
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.value = c
    c1 = ws1.cell(row = 4, column = 9)
    c1.font = f1
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.value = "{:.2f}".format(d)
    c1 = ws1.cell(row = 4, column = 11)
    c1.font = f1
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.value = "{:.2f}".format(c*d)
    if(type=='Sell') : 
        p = stk[[n[0] for n in stk].index(b)][2]
        if(p>d) : f2 = Font(name='Bahnschrift',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FFFF0000')
        else : f2 = Font(name='Bahnschrift',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF008000')
        net = d-p
        c1 = ws1.cell(row = 4, column = 13)
        c1.font = f2
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.value = "{:.2f}".format(net*c)
        c1 = ws1.cell(row = 4, column = 15)
        c1.font = f2
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.value = str("{:.2f}".format(100*net/p))+"%"
        c1 = ws1.cell(row = 4, column = 5)
        c1.font = f2
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.value = type
    else : 
        c1 = ws1.cell(row = 4, column = 13)
        c1.font = f1
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.value = "-"
        c1 = ws1.cell(row = 4, column = 15)
        c1.font = f1
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.value = "-"
    wb.save('Stock.xlsx')

def holiday(t) :
    if(t.weekday() in [5,6]) : return True
    d = int(t.strftime("%d"))
    m = int(t.strftime("%m"))
    y = t.strftime("%Y")
    h21 = [[],[],[],[],[],[],[],[],[19],[10],[15],[4,5,19],[]]
    if(y=='2021') : 
        if(d in h21[m]) : return True
    h22 = [[],[26],[],[1,18],[14,15],[3],[],[],[9,15,31],[],[5,24,25],[8],[]]
    if(y=='2022') : 
        if(d in h22[m]) : return True    
    return False    

def market_open(now) :
    if holiday(now) : return False     
    h = now.hour
    m = now.minute
    if (h>=10 and h<=15 and (m==30 or m==0)) or (h==9 and m==30) : return True
    return False

def report(now) :
    date = now.strftime("%d-%m-%Y")
    wb = load_workbook('dontopen.xlsx')
    ws1 = wb["DayData"]
    nw = ws1.cell(row = 3, column = 3).value
    nw_prev = ws1.cell(row = 4, column = 3).value
    pl = nw-100000
    plp = pl/1000
    pls = ""
    if(pl>0) : pls = "+"
    pl = "{:.2f}".format(pl)
    plp = "{:.2f}".format(plp)
    pl = pls + pl
    plp = pls + plp + " %"
    dpl = nw-nw_prev
    dplp = (100*dpl)/(nw_prev)
    pls = ""
    if(dpl>0) : pls = "+"
    dpl = "{:.2f}".format(dpl)
    dplp = "{:.2f}".format(dplp)
    dpl = pls + dpl
    dplp = pls + dplp + " %"
    s = f"Report  | {date}\n"
    s = s + f"\nP&L     : {pl} ({plp})"
    s = s + f"\nDay P&L : {dpl} ({dplp})"
    return s

client = commands.Bot(command_prefix=["pls ","Pls ","PLS "])

client.remove_command('help')

@client.event
async def on_ready():

    print('Bot is online!')

    @tasks.loop(minutes=30)
    async def chart_poster():
        now = datetime.datetime.now()
        if(now.minute>=30) : mr = 30
        else : mr = 0
        now = now.replace(minute=mr)
        if market_open(now) :
            sc.day(1)
            sc.day(5)
            sc.day(10)
            sc.day(30)
            sc.day(60)
            sc.lpf_day(365)
            await client.get_channel(871450048414777374).purge()
            await client.get_channel(871450048414777374).send(file=discord.File('Stock Charts/Day1.png'))
            await client.get_channel(871450159756750938).purge()
            await client.get_channel(871450159756750938).send(file=discord.File('Stock Charts/Day5.png'))
            await client.get_channel(871450182489894982).purge()
            await client.get_channel(871450182489894982).send(file=discord.File('Stock Charts/Day10.png'))
            await client.get_channel(871450215050260550).purge()
            await client.get_channel(871450215050260550).send(file=discord.File('Stock Charts/Day30.png'))
            await client.get_channel(871450344335482940).purge()
            await client.get_channel(871450344335482940).send(file=discord.File('Stock Charts/Day60.png'))
            await client.get_channel(871450762373386240).purge()
            await client.get_channel(871450762373386240).send(file=discord.File('Stock Charts/Day365.png'))
            if(now.hour==15 and now.minute==30) : await client.get_channel(871454820828065792).send(f"<@375175306161422339>\n```{report(now)}```")
    chart_poster.start()    
    
    @tasks.loop(hours=1)
    async def alert_poster():
        with open('alert.json', 'r') as openfile : alt = json.load(openfile)
        for i in range(len(alt)) : 
            p = price(alt[i][0])
            if(alt[i][2]=='Buy' and p<=alt[i][1] and p>0) : 
                await client.get_channel(860626380931465246).send(f"<@375175306161422339>\n```{alt[i][0]}{' '*(15-len(alt[i][0]))}{alt[i][1]}{' '*(15-len(str(alt[i][1])))}{alt[i][2]}```")
                del alt[i]
                with open("alert.json", "w") as outfile : json.dump(alt, outfile)
                break
            if(alt[i][2]=='Sell' and p>=alt[i][1]) : 
                await client.get_channel(860626424620384276).send(f"<@375175306161422339>\n```{alt[i][0]}{' '*(15-len(alt[i][0]))}{alt[i][1]}{' '*(15-len(str(alt[i][1])))}{alt[i][2]}```")
                del alt[i]
                with open("alert.json", "w") as outfile : json.dump(alt, outfile)
    alert_poster.start()

    @tasks.loop(hours=8)
    async def analysis_updater():
        with open('stocks.json', 'r') as openfile : stk = json.load(openfile)
        update_holdings(stk)
        date = datetime.datetime.now()
        if(date.day==1 and date.hour<=8) : 
            wb = load_workbook('Stock.xlsx')
            ws1 = wb["Holdings"]    
            f1 = Font(name='Arial',size=11,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
            nw = float(ws1.cell(row = 8, column = 14).value)
            ws1 = wb["Analysis"] 
            dm = date.month - 1
            yr = date.year
            if(dm==0) :
                dm = 12
                yr = yr - 1
            month = datetime.date(1900, dm, 1).strftime('%B')
            rw = (4 + (dm-7)*2 + (yr - 2021)*24)
            c1 = ws1.cell(row = rw, column = 1)
            c1.font = f1
            c1.alignment = Alignment(horizontal="center", vertical="center")
            c1.value = f"{month}, {yr}"
            c1 = ws1.cell(row = rw, column = 3)
            c1.font = f1
            c1.alignment = Alignment(horizontal="center", vertical="center")
            c1.value = "{:.2f}".format(nw)
            try : nw_prev =  float(ws1.cell(row = rw-2, column = 3).value)
            except :
                with open('walletdep.json', 'r') as openfile : nw_prev = json.load(openfile)
            clm = 6+dm
            rw = 4 + (yr-2021)*5
            f1_r = Font(name='Calibri',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FFFF0000')
            f1_g = Font(name='Calibri',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF008000')
            f2 = Font(name='Calibri',size=10,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
            f2_r = Font(name='Calibri',size=10,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FFFF0000')
            f2_g = Font(name='Calibri',size=10,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF008000') 
            c1 = ws1.cell(row = rw, column = clm)
            c1.font = f1_g
            if(nw<nw_prev) : c1.font = f1_r
            c1.alignment = Alignment(horizontal="center", vertical="center")
            c1.value = "{:.2f}".format(100*(nw-nw_prev)/nw_prev) + "%"
            c1 = ws1.cell(row = rw+2, column = clm)
            c1.font = f2
            c1.alignment = Alignment(horizontal="center", vertical="center")
            c1.value = "{:.2f}".format(nw)
            c1 = ws1.cell(row = rw+3, column = clm)
            c1.font = f2_g
            if(nw<nw_prev) : c1.font = f2_r
            c1.alignment = Alignment(horizontal="center", vertical="center")
            c1.value = "{:.2f}".format(nw-nw_prev)
            if(date.month==12) : 
                c1 = ws1.cell(row = rw, column = 20)
                c1.font = f1_g
                if(yr==2021) : 
                    with open('walletdep.json', 'r') as openfile : nw_prev = json.load(openfile)
                else : nw_prev = ws1.cell(row = rw-3, column = clm).value
                if(nw<nw_prev) : c1.font = f1_r
                c1.alignment = Alignment(horizontal="center", vertical="center")
                c1.value = "{:.2f}".format(100*(nw-nw_prev)/nw_prev) + "%"   
                c1 = ws1.cell(row = 4 + (yr-2020)*5, column = 5)
                c1.font = Font(name='Calibri',size=11,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
                c1.alignment = Alignment(horizontal="center", vertical="center")
                c1.value = (yr+1)  
                c1 = ws1.cell(row = rw+3, column = 20)
                c1.font = f2_g
                if(yr==2021) : 
                    with open('walletdep.json', 'r') as openfile : nw_prev = json.load(openfile)
                else : nw_prev = ws1.cell(row = rw-3, column = clm).value
                if(nw<nw_prev) : c1.font = f2_r
                c1.alignment = Alignment(horizontal="center", vertical="center")
                c1.value = "{:.2f}".format(nw-nw_prev) + "%"  
                c1 = ws1.cell(row = rw+2, column = 20)
                c1.font = f2
                c1.alignment = Alignment(horizontal="center", vertical="center")
                c1.value = "{:.2f}".format(nw)  
                
            wb.save('Stock.xlsx')    
    analysis_updater.start() 
    
@client.command(pass_context = True)
async def help(ctx):
    embed = discord.Embed(colour = discord.Colour.dark_green())
    embed.set_author(name = 'Stock Portfolio Organiser')
    embed.add_field(name='Stock Price',value = 'pls price <NSE Code>', inline = False)
    embed.add_field(name='Net Worth',value = 'pls bal', inline = False) 
    embed.add_field(name='Alerts',value = 'pls add <NSE Code> <Price>\npls alerts\npls remove', inline = False)
    embed.add_field(name='Invest',value = 'pls buy <NSE Code> <Quantity> <Price>\npls sell <NSE Code> <Quantity> <Price> ', inline = False)
    embed.add_field(name='Wallet',value = 'pls deposit\npls withdraw\npls brockerage', inline = False)
    embed.add_field(name='Reset Json Files',value = 'pls clear', inline = False)
    await ctx.send(embed=embed) 

@client.command(aliases = ["price","PRICE","p","P","prc","Prc"])
async def Price(ctx,*,a = None):
    
    if a is None : 
        await ctx.send("```Enter Stock Code - ```")            
        msg = await client.wait_for('message', check=lambda message: message.author == ctx.author)
        a = msg.content
    
    await ctx.send(f"```{str(price(a))}```")

@client.command(aliases = ["Add","alert","Alert"])
async def add(ctx,*,a = None):   

    if a is None : 
        await ctx.send("```Enter Stock Code - ```")            
        msg = await client.wait_for('message', check=lambda message: message.author == ctx.author)
        b = (msg.content).upper()
        await ctx.send("```Enter Target Price - ```")            
        msg = await client.wait_for('message', check=lambda message: message.author == ctx.author)
        c = float(msg.content)
    else :   
        for i in range(len(a)) : 
            if (a[i]==' ') : 
                b = (a[:i]).upper()
                c = float(a[i+1:])
                break
            if(i==len(a)-1) : 
                b = a.upper()
                await ctx.send("```Enter Target Price - ```")            
                msg = await client.wait_for('message', check=lambda message: message.author == ctx.author)
                c = float(msg.content)
                
    try :
        with open('alert.json', 'r') as openfile : alt = json.load(openfile)
    except : alt = []

    exist = True
    prc = price(b)
    if(prc==-1) : exist = False 
    if exist : 
        d = "Buy" 
        if(prc<c) : d = "Sell" 
        alt.append([b,c,d])
        with open("alert.json", "w") as outfile : json.dump(alt, outfile)
        await ctx.send(f"```Added Alert - {[b,c,d]} ```") 
    else : await ctx.send(f"```No Such Stock```")   

@client.command(aliases = ["Alerts"])
async def alerts(ctx,*,a = None):  
    try :
        with open('alert.json', 'r') as openfile : alt = json.load(openfile)
    except : alt = []
    s = ""
    for i in range(len(alt)) : s = s + f"{i+1}){' '*(len(str(len(alt)))-len(str(i+1)))} {alt[i][0]}{' '*(15-len(alt[i][0]))}{alt[i][1]}{' '*(15-len(str(alt[i][1])))}{alt[i][2]}\n"
    await ctx.send(f"```{s}```") 

@client.command(aliases = ["Remove","Rem","rem"])
async def remove(ctx,*,a = None):  

    try :
        with open('alert.json', 'r') as openfile : alt = json.load(openfile)
    except : alt = []

    if a is None : 
        s = ""
        for i in range(len(alt)) : s = s + f"{i+1}){' '*(len(str(len(alt)))-len(str(i+1)))} {alt[i][0]}{' '*(15-len(alt[i][0]))}{alt[i][1]}{' '*(15-len(str(alt[i][1])))}{alt[i][2]}\n"
        await ctx.send(f"```{s}```") 
        await ctx.send("```Enter Index of Alerts to Remove - ```")            
        msg = await client.wait_for('message', check=lambda message: message.author == ctx.author)
        a = msg.content

    b = a.split()
    for i in range(len(b)) : b[i] = int(b[i])-1
    alt = [i for j, i in enumerate(alt) if j not in b]
    with open("alert.json", "w") as outfile : json.dump(alt, outfile)
    await ctx.send(f"```Removed```")   

@client.command(aliases = ["Buy"])
async def buy(ctx,*,a = None):  

    try :
        with open('stocks.json', 'r') as openfile : stk = json.load(openfile)
    except : stk = []

    if a is None : 
        await ctx.send("```Enter Stock Code - ```")            
        msg = await client.wait_for('message', check=lambda message: message.author == ctx.author)
        b = (msg.content).upper()
        await ctx.send("```Enter Quantity - ```")            
        msg = await client.wait_for('message', check=lambda message: message.author == ctx.author)
        c = int(msg.content)
        await ctx.send("```Enter Price - ```")            
        msg = await client.wait_for('message', check=lambda message: message.author == ctx.author)
        d = float(msg.content)
    else :   
        for i in range(len(a)) : 
            if (a[i]==' ') : 
                b = (a[:i]).upper()
                for j in range(i+1,len(a)) : 
                    if (a[j]==' ') :
                        c = int(a[i+1:j])
                        d = float(a[j+1:])
                        break
                break  

    prc = price(b)
    if(prc!=-1) : 
        if b not in [n[0] for n in stk] : stk.append([b,c,d])
        else : 
            ind = [n[0] for n in stk].index(b)
            p = (c*d + stk[ind][1]*stk[ind][2])/(c + stk[ind][1])
            stk[ind][1] = stk[ind][1] + c
            stk[ind][2] = p
        stk = sorted(stk, key=lambda x:(x[1]*x[2]), reverse = True)  
        with open("stocks.json", "w") as outfile : json.dump(stk, outfile)
        update_holdings(stk)
        update_wallet(c*d,"Stock Buy")
        update_trades(stk,b,c,d,"Buy")
        await ctx.send(f"```Brought Stock - {[b,c,d,'{:.2f}'.format(c*d)]} ```") 
    else : await ctx.send(f"```No Such Stock```")                

@client.command(aliases = ["Sell"])
async def sell(ctx,*,a = None):  

    try :
        with open('stocks.json', 'r') as openfile : stk = json.load(openfile)
    except : stk = []

    if a is None : 
        await ctx.send("```Enter Stock Code - ```")            
        msg = await client.wait_for('message', check=lambda message: message.author == ctx.author)
        b = (msg.content).upper()
        await ctx.send("```Enter Quantity - ```")            
        msg = await client.wait_for('message', check=lambda message: message.author == ctx.author)
        c = int(msg.content)
        await ctx.send("```Enter Price - ```")            
        msg = await client.wait_for('message', check=lambda message: message.author == ctx.author)
        d = float(msg.content)
    else :   
        for i in range(len(a)) : 
            if (a[i]==' ') : 
                b = (a[:i]).upper()
                for j in range(i+1,len(a)) : 
                    if (a[j]==' ') :
                        c = int(a[i+1:j])
                        d = float(a[j+1:])
                        break
                break  

    prc = price(b)
    if(prc!=-1) :
        update_trades(stk,b,c,d,"Sell")
        ind = [n[0] for n in stk].index(b)
        if (stk[ind][1]==c) : del stk[ind]
        else : stk[ind][1] = stk[ind][1] - c 
        with open("stocks.json", "w") as outfile : json.dump(stk, outfile)
        update_holdings(stk)
        update_wallet(c*d,"Stock Sell")
        await ctx.send(f"```Sold Stock - {[b,c,d,'{:.2f}'.format(c*d)]} ```") 
    else : await ctx.send(f"```No Such Stock```")


@client.command(aliases = ["Bal","balance","Balance"])
async def bal(ctx,*,a=None): 
    with open("wallbal.json", "r") as openfile : w = json.load(openfile)
    with open("stocks.json", "r") as openfile : stk = json.load(openfile)
    with open('walletdep.json', 'r') as openfile : wdep = json.load(openfile)
    sum = 0 
    s = ""
    if a is not None : 
        pl = []
        a = a.split()
        if len(a)==1 :
            for i in range(len(stk)) :
                p = price(stk[i][0])
                pl.append((p-stk[i][2])*stk[i][1])
        if len(a)>1 :
            for i in range(len(stk)) :
                p = price(stk[i][0])
                pl.append((p-stk[i][2])/stk[i][2])        
        stk = [x for _,x in sorted(zip(pl,stk),reverse=True)]

    for i in range(len(stk)) :  
        p = price(stk[i][0])
        sum = sum + p*stk[i][1]
        s2 = ""
        s5 = ""
        if(p>=stk[i][2]) : 
            s2 = "+"
            s5 = "+"
        s2 = s2 + "{:.2f}".format(100*(p-stk[i][2])/stk[i][2]) + " %"
        s3 = "{:.2f}".format(p)
        s4 = "{:.2f}".format(p*stk[i][1])
        s5 = s5 + "{:.2f}".format((p-stk[i][2])*stk[i][1])
        s = s + f"{i+1}){' '*(len(str(len(stk)))-len(str(i+1)))} {stk[i][0]}{' '*(15-len(stk[i][0]))}{stk[i][1]}{' '*(8-len(str(stk[i][1])))}{s3}{' '*(13-len(s3))}{s4}{' '*(15-len(s4))}|   {s5}{' '*(15-len(s5))}{s2}\n"
    net = w + sum
    s1 = ""
    if(net>=wdep) : s1 = "+"
    try :
        s1 = s1 + "{:.2f}".format(100*(net-wdep)/wdep)
        s1 = s1[0] + ' ' + s1[1:]
    except : s1 = ""
    nets = "{:.2f}".format(net)
    ws = "{:.2f}".format(w)
    sum = "{:.2f}".format(sum)
    now = datetime.datetime.now()
    dt1 = now.strftime("%d-%m-%Y")
    dt2 = now.strftime("%I:%M %p")
    dt = dt1 + ' '*(15-len(dt1)) + dt2
    nw = f"Net Worth : {nets} ({s1} %)"
    sp = ' '*(55-len(nw))
    s = f"{nw}{sp}|   {dt}\n----------------------------------------------\nStocks : {sum}   |   Wallet : {ws}\n\n" + s
    await ctx.send(f"```{s}```")

@client.command(aliases = ["Deposit","dep","Dep"])
async def deposit(ctx,*,a = None): 
    if a is None : 
        await ctx.send("```Enter Amount - ```")            
        msg = await client.wait_for('message', check=lambda message: message.author == ctx.author)
        a = float(msg.content) 
    else : a = float(a)       
    update_wallet(a,"Deposit")  
    await ctx.send(f"```Deposited {a}```")

@client.command(aliases = ["Withdraw","With","with"])
async def withdraw(ctx,*,a = None):  
    if a is None : 
        await ctx.send("```Enter Amount - ```")            
        msg = await client.wait_for('message', check=lambda message: message.author == ctx.author)
        a = float(msg.content) 
    else : a = float(a)          
    update_wallet(a,"Withdraw") 
    await ctx.send(f"```Withdrew {a}```")

@client.command(aliases = ["brock","Brock","Brockerage","tax","Tax"])
async def brockerage(ctx,*,a = None):   
    if a is None : 
        await ctx.send("```Enter Amount - ```")            
        msg = await client.wait_for('message', check=lambda message: message.author == ctx.author)
        a = float(msg.content) 
    else : a = float(a)                    
    update_wallet(a,"Brockerage") 
    with open('tax.json', 'r') as openfile : tax = json.load(openfile)
    tax = tax + a
    with open("tax.json", "w") as outfile : json.dump(tax, outfile)
    await ctx.send(f"```Updated```") 

@client.command(aliases = ["Cost","costs","Costs","fee","Fee","fees","Fees"])
async def cost(ctx):      
    with open('tax.json', 'r') as openfile : tax = json.load(openfile)  
    await ctx.send(f"```Tax : {tax}```") 

@client.command(aliases = ["Rep","report","Report"])
async def rep(ctx):      
    await ctx.send(f"```{report(datetime.datetime.now())}```")          

@client.command(aliases = ["Chart"])
async def chart(ctx,*,a=None):      
    if a is None : a = sc.max()
    a = int(a)
    sc.day(a)
    await ctx.send(file=discord.File(f'Stock Charts/Day{a}.png'))      

@client.command(aliases = ["Clear","clr","Clr"])
async def clear(ctx,*,a = None):
    a = []
    with open("stocks.json", "w") as outfile : json.dump(a, outfile)
    a = [0,0]
    with open("traderow.json", "w") as outfile : json.dump(a, outfile)
    a = 0.0
    with open("wallbal.json", "w") as outfile : json.dump(a, outfile)
    with open("walletdep.json", "w") as outfile : json.dump(a, outfile)
    await ctx.send(f"```Cleared```")


client.run("ODYwNjU5MTUwMjM3MjA0NTQw.YN-dSw.GhfqfeicY01VTCQwFd6T82tHiXY")